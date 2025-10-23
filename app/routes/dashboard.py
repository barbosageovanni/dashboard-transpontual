# app/routes/dashboard.py
from flask import Blueprint, render_template, jsonify, request
from flask_login import login_required, current_user
from app.models.cte import CTE
from app.models.permissions import PermissionManager
from app import db
from datetime import datetime, timedelta
import pandas as pd
import sys
import os

# Mantido: compat para serviços opcionais
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'services'))
try:
    from app.services.metricas_service import MetricasService
    METRICAS_SERVICE_OK = True
except Exception as e:
    print(f"[AVISO] MetricasService nao disponivel: {e}")
    METRICAS_SERVICE_OK = False

bp = Blueprint('dashboard', __name__, url_prefix='/dashboard')

@bp.route('/')
@login_required
def index():
    # Obter módulos que o usuário tem acesso
    user_modules = PermissionManager.get_modules_for_display(current_user.id)

    # Verificar se tem acesso aos módulos específicos
    can_access_financial = any(m['key'] == 'financeiro' for m in user_modules)
    can_access_fleet = any(m['key'] == 'frotas' for m in user_modules)
    can_access_admin = any(m['key'] == 'admin' for m in user_modules)

    # Verificar integração com sistema de frotas
    integration_available = True  # Assumir que está disponível por enquanto
    frotas_system_url = "http://localhost:8050"  # URL do sistema de frotas
    is_jwt_authenticated = False  # Por enquanto
    frotas_user_info = None  # Por enquanto

    return render_template('dashboard/frotas_style.html',
                         user_modules=user_modules,
                         can_access_financial=can_access_financial,
                         can_access_fleet=can_access_fleet,
                         can_access_admin=can_access_admin,
                         integration_available=integration_available,
                         frotas_system_url=frotas_system_url,
                         is_jwt_authenticated=is_jwt_authenticated,
                         frotas_user_info=frotas_user_info)

@bp.route('/api/debug')
@login_required
def api_debug():
    """Diagnóstico rápido do backend e conversão"""
    try:
        total_ctes = CTE.query.count()
        exemplos = CTE.query.limit(5).all()
        exemplos_dict = []
        for cte in exemplos:
            try:
                exemplos_dict.append({
                    'numero_cte': cte.numero_cte,
                    'valor_total': float(cte.valor_total or 0),
                    'destinatario_nome': cte.destinatario_nome,
                    'has_baixa': bool(cte.data_baixa is not None),
                    'data_emissao': cte.data_emissao.isoformat() if cte.data_emissao else None,
                    'data_atesto': cte.data_atesto.isoformat() if cte.data_atesto else None
                })
            except Exception as e:
                exemplos_dict.append({'erro': str(e)})
        return jsonify({
            'success': True,
            'debug': {
                'total_ctes_db': total_ctes,
                'exemplos': exemplos_dict,
                'metricas_service_available': METRICAS_SERVICE_OK,
                'timestamp': datetime.now().isoformat()
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/metricas')
@login_required
def api_metricas():
    """API para métricas do dashboard principal"""
    try:
        # Obter estatísticas básicas dos CTEs
        total_ctes = CTE.query.count()

        # Receita total
        receita_total = db.session.query(db.func.sum(CTE.valor_total)).scalar() or 0

        # Processos completos (CTEs com envio_final preenchido)
        processos_completos = CTE.query.filter(CTE.envio_final.isnot(None)).count()

        # Alertas pendentes (CTEs sem primeiro_envio há mais de 30 dias)
        from datetime import datetime, timedelta
        data_limite = datetime.now() - timedelta(days=30)
        alertas_pendentes = CTE.query.filter(
            CTE.primeiro_envio.is_(None),
            CTE.data_emissao < data_limite
        ).count()

        return jsonify({
            'success': True,
            'metricas': {
                'total_ctes': total_ctes,
                'receita_total': float(receita_total),
                'processos_completos': processos_completos,
                'alertas_pendentes': alertas_pendentes
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/api/dashboard/metricas')
@login_required
def api_dashboard_metricas():
    """Métricas principais do dashboard (corrigidas)"""
    try:
        return _calcular_metricas_completas()
    except Exception as e:
        print(f"[ERROR] Erro na API de métricas: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'metricas': _metricas_vazias(),
            'alertas': {},
            'variacoes': {},
            'graficos': _graficos_vazios()
        }), 500

def _carregar_df_cte() -> pd.DataFrame:
    """
    🔧 FUNÇÃO CORRIGIDA - Carrega dados usando SQL string ao invés de SQLAlchemy statement
    """
    try:
        # ✅ CORREÇÃO PRINCIPAL: Usar SQL string diretamente
        sql_query = """
            SELECT numero_cte, destinatario_nome, veiculo_placa, valor_total,
                   data_emissao, numero_fatura, data_baixa, observacao,
                   data_inclusao_fatura, data_envio_processo, primeiro_envio,
                   data_rq_tmc, data_atesto, envio_final, origem_dados
            FROM dashboard_baker 
            ORDER BY numero_cte DESC
        """
        
        # ✅ Usando connection() ao invés de bind para evitar warnings
        df = pd.read_sql_query(sql_query, db.engine)
        
        # Consertos de tipo
        if 'valor_total' in df.columns:
            df['valor_total'] = pd.to_numeric(df['valor_total'], errors='coerce').fillna(0)

        date_cols = [
            'data_emissao', 'data_baixa', 'data_inclusao_fatura', 'data_envio_processo',
            'primeiro_envio', 'data_rq_tmc', 'data_atesto', 'envio_final'
        ]
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')

        print(f"[OK] DataFrame carregado: {len(df)} registros")
        return df
        
    except Exception as e:
        print(f"[ERRO] Erro ao carregar DataFrame: {e}")
        # Fallback: retornar DataFrame vazio com colunas necessárias
        return pd.DataFrame(columns=[
            'numero_cte', 'destinatario_nome', 'veiculo_placa', 'valor_total',
            'data_emissao', 'numero_fatura', 'data_baixa', 'observacao',
            'data_inclusao_fatura', 'data_envio_processo', 'primeiro_envio',
            'data_rq_tmc', 'data_atesto', 'envio_final', 'origem_dados'
        ])

def _calcular_metricas_completas():
    """Calcula métricas, alertas, variações e dados para gráficos."""
    df = _carregar_df_cte()

    if df.empty:
        print("[WARN] DataFrame vazio - retornando métricas zeradas")
        return jsonify({
            'success': True,
            'metricas': _metricas_vazias(),
            'alertas': {},
            'variacoes': {},
            'graficos': _graficos_vazios(),
            'timestamp': datetime.now().isoformat(),
            'total_registros': 0
        })

    print(f"[CALC] Calculando métricas para {len(df)} registros")
    metricas = _metricas_basicas(df)
    alertas = calcular_alertas_inteligentes(df)  # ✅ CORRIGIDO!
    variacoes = _variacoes(df)
    graficos = _graficos(df)

    return jsonify({
        'success': True,
        'metricas': metricas,
        'alertas': alertas,
        'variacoes': variacoes,
        'graficos': graficos,
        'timestamp': datetime.now().isoformat(),
        'total_registros': int(len(df))
    })

def _metricas_basicas(df: pd.DataFrame) -> dict:
    """Calcula métricas básicas com tratamento de erros robusto"""
    if df.empty:
        return _metricas_vazias()

    try:
        total_ctes = int(len(df))
        clientes_unicos = int(df['destinatario_nome'].nunique()) if 'destinatario_nome' in df.columns else 0
        veiculos_ativos = int(df['veiculo_placa'].nunique()) if 'veiculo_placa' in df.columns else 0

        # Garantir que valor_total seja numérico
        valores = pd.to_numeric(df['valor_total'], errors='coerce').fillna(0)
        valor_total = float(valores.sum())

        # Métricas de baixa
        has_baixa = df['data_baixa'].notna() if 'data_baixa' in df.columns else pd.Series([False] * len(df))
        f_pagas = int(has_baixa.sum())
        f_pend = int((~has_baixa).sum())
        valor_pago = float(valores[has_baixa].sum()) if any(has_baixa) else 0.0
        valor_pend = float(valores[~has_baixa].sum()) if any(~has_baixa) else 0.0

        # Métricas de fatura
        tem_fatura = False
        if 'numero_fatura' in df.columns:
            tem_fatura = df['numero_fatura'].notna() & (df['numero_fatura'] != '')
        
        ctes_com_fatura = int(tem_fatura.sum()) if hasattr(tem_fatura, 'sum') else 0
        ctes_sem_fatura = int(total_ctes - ctes_com_fatura)
        valor_com_fatura = float(valores[tem_fatura].sum()) if hasattr(tem_fatura, 'sum') and any(tem_fatura) else 0.0
        valor_sem_fatura = float(valor_total - valor_com_fatura)

        # Processos completos
        proc_completos = 0
        if all(col in df.columns for col in ['data_emissao', 'primeiro_envio', 'data_atesto', 'envio_final']):
            proc_completos_mask = (
                df['data_emissao'].notna() &
                df['primeiro_envio'].notna() &
                df['data_atesto'].notna() &
                df['envio_final'].notna()
            )
            proc_completos = int(proc_completos_mask.sum())
        
        proc_incompletos = int(total_ctes - proc_completos)

        # Estatísticas de valor
        ticket_medio = float(valores.mean()) if total_ctes > 0 else 0.0
        maior_valor = float(valores.max()) if total_ctes > 0 else 0.0
        menor_valor = float(valores.min()) if total_ctes > 0 else 0.0

        # Receita mensal e crescimento
        receita_mensal_media = 0.0
        crescimento_mensal = 0.0
        if 'data_emissao' in df.columns and df['data_emissao'].notna().any():
            try:
                tmp = df[df['data_emissao'].notna()].copy()
                tmp['valor_numerico'] = valores[df['data_emissao'].notna()]
                tmp['mes_ano'] = tmp['data_emissao'].dt.to_period('M')
                receita = tmp.groupby('mes_ano')['valor_numerico'].sum()
                if len(receita) > 0:
                    receita_mensal_media = float(receita.mean())
                    if len(receita) >= 2 and float(receita.iloc[-2]) > 0:
                        crescimento_mensal = float((receita.iloc[-1] - receita.iloc[-2]) / receita.iloc[-2] * 100)
            except Exception as e:
                print(f"[WARN] Erro no cálculo de receita mensal: {e}")

        return {
            'total_ctes': total_ctes,
            'clientes_unicos': clientes_unicos,
            'veiculos_ativos': veiculos_ativos,
            'valor_total': valor_total,
            'valor_pago': valor_pago,
            'valor_pendente': valor_pend,
            'faturas_pagas': f_pagas,
            'faturas_pendentes': f_pend,
            'ctes_com_fatura': ctes_com_fatura,
            'ctes_sem_fatura': ctes_sem_fatura,
            'valor_com_fatura': valor_com_fatura,
            'valor_sem_fatura': valor_sem_fatura,
            'processos_completos': proc_completos,
            'processos_incompletos': proc_incompletos,
            'ticket_medio': ticket_medio,
            'maior_valor': maior_valor,
            'menor_valor': menor_valor,
            'receita_mensal_media': receita_mensal_media,
            'crescimento_mensal': crescimento_mensal,
            'taxa_conclusao': float(proc_completos / total_ctes * 100) if total_ctes else 0.0,
            'taxa_pagamento': float(f_pagas / total_ctes * 100) if total_ctes else 0.0,
            'taxa_faturamento': float(ctes_com_fatura / total_ctes * 100) if total_ctes else 0.0
        }
        
    except Exception as e:
        print(f"[ERROR] Erro no cálculo de métricas básicas: {e}")
        return _metricas_vazias()

def _alertas(df: pd.DataFrame) -> dict:
    """Calcula alertas inteligentes com tratamento robusto de erros"""
    alertas = {
        'primeiro_envio_pendente': {'qtd': 0, 'valor': 0.0, 'lista': []},
        'envio_final_pendente': {'qtd': 0, 'valor': 0.0, 'lista': []},
        'faturas_vencidas': {'qtd': 0, 'valor': 0.0, 'lista': []},
        'ctes_sem_faturas': {'qtd': 0, 'valor': 0.0, 'lista': []}
    }
    
    if df.empty:
        return alertas
    
    try:
        hoje = pd.Timestamp.now().normalize()
        valores = pd.to_numeric(df['valor_total'], errors='coerce').fillna(0)

        def _safe_item(row, extra_field):
            """Cria item seguro para lista de alertas"""
            try:
                base = {
                    'numero_cte': int(row['numero_cte']) if pd.notna(row['numero_cte']) else 0,
                    'destinatario_nome': str(row['destinatario_nome']) if pd.notna(row['destinatario_nome']) else 'N/A',
                    'valor_total': float(valores[row.name]) if pd.notna(valores[row.name]) else 0.0
                }
                if extra_field in row and pd.notna(row[extra_field]):
                    if hasattr(row[extra_field], 'isoformat'):
                        base[extra_field] = row[extra_field].isoformat()
                    else:
                        base[extra_field] = str(row[extra_field])
                else:
                    base[extra_field] = None
                return base
            except Exception as e:
                print(f"[WARN] Erro ao criar item de alerta: {e}")
                return {'numero_cte': 0, 'destinatario_nome': 'Erro', 'valor_total': 0.0, extra_field: None}

        # 1) Primeiro envio pendente (>1 dias após emissão)
        if all(col in df.columns for col in ['data_emissao', 'primeiro_envio']):
            mask = (df['data_emissao'].notna() &
                    ((hoje - df['data_emissao']).dt.days > 10) &
                    df['primeiro_envio'].isna())
            if mask.any():
                c = df[mask]
                alertas['primeiro_envio_pendente'] = {
                    'qtd': int(len(c)),
                    'valor': float(valores[mask].sum()),
                    'lista': [_safe_item(r, 'data_emissao') for _, r in c.iterrows()]
                }

        # 2) Envio final pendente (>1 dias após atesto)
        if all(col in df.columns for col in ['data_atesto', 'envio_final']):
            mask_envio_final = (
                df['data_atesto'].notna() & 
                ((hoje - df['data_atesto']).dt.days > 1) &  # MUDAR DE 5 PARA 1
                df['envio_final'].isna()
            )
            if mask_envio_final.any():
                c = df[mask_envio_final]
                alertas['envio_final_pendente'] = {
                    'qtd': int(len(c)),
                    'valor': float(valores[mask_envio_final].sum()),
                    'lista': [_safe_item(r, 'data_atesto') for _, r in c.iterrows()]
                }

        # 3) Faturas vencidas (>90 dias do atesto, sem baixa)
        if all(col in df.columns for col in ['data_atesto', 'data_baixa']):
            mask_vencidas = (
                    df['envio_final'].notna() &  # MUDAR DE data_atesto PARA envio_final
                    ((hoje - df['envio_final']).dt.days > 90) &
                    df['data_baixa'].isna()
                )
            if mask_vencidas.any():
                c = df[mask_vencidas]
                alertas['faturas_vencidas'] = {
                    'qtd': int(len(c)),
                    'valor': float(valores[mask_vencidas].sum()),
                    'lista': [_safe_item(r, 'data_atesto') for _, r in c.iterrows()]
                }

        # 4) CTEs sem fatura (>3 dias do atesto)
        if all(col in df.columns for col in ['data_atesto', 'numero_fatura']):
            mask = (df['data_atesto'].notna() &
                    ((hoje - df['data_atesto']).dt.days > 3) &
                    (df['numero_fatura'].isna() | (df['numero_fatura'] == '')))
            if mask.any():
                c = df[mask]
                alertas['ctes_sem_faturas'] = {
                    'qtd': int(len(c)),
                    'valor': float(valores[mask].sum()),
                    'lista': [_safe_item(r, 'data_atesto') for _, r in c.iterrows()]
                }

    except Exception as e:
        print(f"[WARN] Erro no cálculo de alertas: {e}")

    return alertas

def _variacoes(df: pd.DataFrame) -> dict:
    """Calcula variações de tempo entre processos"""
    configs = [
        ('rq_tmc_primeiro_envio', 'RQ/TMC - 1º Envio', 'data_rq_tmc', 'primeiro_envio', 3),
        ('primeiro_envio_atesto', '1º Envio - Atesto', 'primeiro_envio', 'data_atesto', 7),
        ('atesto_envio_final', 'Atesto - Envio Final', 'data_atesto', 'envio_final', 2),
        ('cte_inclusao_fatura', 'CTE - Inclusão Fatura', 'data_emissao', 'data_inclusao_fatura', 2),
        ('cte_baixa', 'CTE - Baixa', 'data_emissao', 'data_baixa', 30),
    ]
    
    out = {}
    if df.empty:
        return out

    try:
        for code, nome, inicio, fim, meta in configs:
            if inicio in df.columns and fim in df.columns:
                m = df[inicio].notna() & df[fim].notna()
                if not m.any():
                    continue
                
                _dif = (df.loc[m, fim] - df.loc[m, inicio]).dt.days
                _dif = _dif[_dif >= 0]  # Apenas diferenças positivas
                
                if len(_dif) == 0:
                    continue
                
                media = float(_dif.mean())
                mediana = float(_dif.median())
                
                # Determinar performance
                if media <= meta:
                    perf = 'excelente'
                elif media <= meta * 1.5:
                    perf = 'bom'
                elif media <= meta * 2:
                    perf = 'atencao'
                else:
                    perf = 'critico'
                
                out[code] = {
                    'nome': nome,
                    'media': round(media, 1),
                    'mediana': round(mediana, 1),
                    'qtd': int(len(_dif)),
                    'meta_dias': int(meta),
                    'performance': perf,
                    'min': int(_dif.min()),
                    'max': int(_dif.max())
                }
                
    except Exception as e:
        print(f"[WARN] Erro no cálculo de variações: {e}")

    return out

def _graficos(df: pd.DataFrame) -> dict:
    """Gera dados para gráficos do dashboard"""
    graficos = {
        'evolucao_mensal': {'labels': [], 'valores': [], 'quantidades': []},
        'top_clientes': {'labels': [], 'valores': []},
        'distribuicao_status': {
            'baixas': {'labels': [], 'valores': []},
            'processos': {'labels': [], 'valores': []}
        },
        'performance_veiculos': {'labels': [], 'valores': [], 'quantidades': []}
    }
    
    if df.empty:
        return graficos

    try:
        valores = pd.to_numeric(df['valor_total'], errors='coerce').fillna(0)

        # Evolução mensal
        if 'data_emissao' in df.columns and df['data_emissao'].notna().any():
            try:
                tmp = df[df['data_emissao'].notna() & valores.notna()].copy()
                if not tmp.empty:
                    tmp['mes_ano'] = tmp['data_emissao'].dt.to_period('M')
                    receita = tmp.groupby('mes_ano').agg(
                        valor_total=('valor_total', lambda x: pd.to_numeric(x, errors='coerce').sum()),
                        qtd=('numero_cte', 'count')
                    ).reset_index().tail(12)
                    
                    if not receita.empty:
                        graficos['evolucao_mensal'] = {
                            'labels': [str(p) for p in receita['mes_ano']],
                            'valores': [float(v) for v in receita['valor_total']],
                            'quantidades': [int(q) for q in receita['qtd']]
                        }
            except Exception as e:
                print(f"[WARN] Erro na evolução mensal: {e}")

        # Top clientes
        if 'destinatario_nome' in df.columns:
            try:
                tmp_df = df.copy()
                tmp_df['valor_total'] = valores
                top = tmp_df.groupby('destinatario_nome')['valor_total'].sum().sort_values(ascending=False).head(10)
                if not top.empty:
                    graficos['top_clientes'] = {
                        'labels': [str(i) for i in top.index],
                        'valores': [float(v) for v in top.values]
                    }
            except Exception as e:
                print(f"[WARN] Erro no top clientes: {e}")

        # Distribuição de status
        try:
            com_baixa = int(df['data_baixa'].notna().sum()) if 'data_baixa' in df.columns else 0
            sem_baixa = int(len(df) - com_baixa)
            
            proc_compl = 0
            if all(col in df.columns for col in ['data_emissao', 'primeiro_envio', 'data_atesto', 'envio_final']):
                proc_compl = int((df['data_emissao'].notna() & df['primeiro_envio'].notna() &
                                  df['data_atesto'].notna() & df['envio_final'].notna()).sum())
            proc_incompl = int(len(df) - proc_compl)
            
            graficos['distribuicao_status'] = {
                'baixas': {'labels': ['Com Baixa', 'Sem Baixa'], 'valores': [com_baixa, sem_baixa]},
                'processos': {'labels': ['Completos', 'Incompletos'], 'valores': [proc_compl, proc_incompl]}
            }
        except Exception as e:
            print(f"[WARN] Erro na distribuição de status: {e}")

        # Performance de veículos
        if 'veiculo_placa' in df.columns:
            try:
                tmp_df = df.copy()
                tmp_df['valor_total'] = valores
                v = tmp_df.groupby('veiculo_placa').agg(
                    valor_total=('valor_total', 'sum'),
                    qtd=('numero_cte', 'count')
                ).sort_values('valor_total', ascending=False).head(10)
                
                if not v.empty:
                    graficos['performance_veiculos'] = {
                        'labels': [str(i) for i in v.index],
                        'valores': [float(x) for x in v['valor_total']],
                        'quantidades': [int(x) for x in v['qtd']]
                    }
            except Exception as e:
                print(f"[WARN] Erro na performance de veículos: {e}")

    except Exception as e:
        print(f"[WARN] Erro geral nos gráficos: {e}")

    return graficos

def _metricas_vazias():
    """Retorna estrutura de métricas zeradas"""
    return {
        'total_ctes': 0, 'clientes_unicos': 0, 'valor_total': 0.0,
        'faturas_pagas': 0, 'faturas_pendentes': 0, 'valor_pago': 0.0,
        'valor_pendente': 0.0, 'veiculos_ativos': 0,
        'processos_completos': 0, 'processos_incompletos': 0,
        'ticket_medio': 0.0, 'maior_valor': 0.0, 'menor_valor': 0.0,
        'receita_mensal_media': 0.0, 'crescimento_mensal': 0.0,
        'taxa_conclusao': 0.0, 'taxa_pagamento': 0.0, 'taxa_faturamento': 0.0
    }

def _graficos_vazios():
    """Retorna estrutura de gráficos vazia"""
    return {
        'evolucao_mensal': {'labels': [], 'valores': [], 'quantidades': []},
        'top_clientes': {'labels': [], 'valores': []},
        'distribuicao_status': {
            'baixas': {'labels': [], 'valores': []},
            'processos': {'labels': [], 'valores': []}
        },
        'performance_veiculos': {'labels': [], 'valores': [], 'quantidades': []}
    }

@bp.route('/api/graficos')
@login_required
def api_graficos():
    """API de gráficos - retorna mesma estrutura de métricas"""
    return api_metricas()

@bp.route('/relatorios')
@login_required
def relatorios():
    return render_template('dashboard/relatorios.html')

@bp.route('/api/relatorio/executivo')
@login_required
def api_relatorio_executivo():
    """Gera relatório executivo baseado nas métricas"""
    try:
        data = _calcular_metricas_completas().get_json()
        if not data.get('success'):
            raise Exception(data.get('error', 'Erro desconhecido'))

        m = data['metricas']
        a = data['alertas']
        v = data.get('variacoes', {})

        relatorio = {
            'titulo': 'Relatório Executivo - Dashboard Baker',
            'data_geracao': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'resumo_financeiro': {
                'total_ctes': m['total_ctes'],
                'valor_total': m['valor_total'],
                'receita_realizada': m['valor_pago'],
                'valor_pendente': m['valor_pendente'],
                'ticket_medio': m['ticket_medio'],
                'crescimento_mensal': m['crescimento_mensal']
            },
            'indicadores_chave': {
                'taxa_conclusao': m['taxa_conclusao'],
                'taxa_pagamento': m['taxa_pagamento'],
                'taxa_faturamento': m['taxa_faturamento'],
                'clientes_ativos': m['clientes_unicos'],
                'veiculos_ativos': m['veiculos_ativos']
            },
            'alertas_criticos': {
                'total_alertas': sum(x.get('qtd', 0) for x in a.values()),
                'valor_em_risco': sum(x.get('valor', 0.0) for x in a.values()),
                'detalhes': a
            },
            'performance_temporal': v
        }

        return jsonify({'success': True, 'relatorio': relatorio})

    except Exception as e:
        print(f"[ERROR] Erro no relatório executivo: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ================================
# ROTAS PARA VALORES PENDENTES
# ================================

@bp.route('/api/valores-pendentes')
@login_required
def api_valores_pendentes():
    """API para listar CTEs com valores pendentes (sem baixa)"""
    try:
        # Parâmetros de paginação e filtro
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        cliente_filtro = request.args.get('cliente', '').strip()

        # Query base: CTEs sem baixa
        query = CTE.query.filter(CTE.data_baixa.is_(None))

        # Aplicar filtro de cliente se fornecido
        if cliente_filtro:
            query = query.filter(CTE.destinatario_nome.ilike(f'%{cliente_filtro}%'))

        # Ordenar por data de emissão (mais antigos primeiro)
        query = query.order_by(CTE.data_emissao.asc())

        # Executar paginação
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        ctes = pagination.items

        # Preparar dados para resposta
        lista_ctes = []
        total_valor = 0.0

        for cte in ctes:
            valor = float(cte.valor_total or 0)
            total_valor += valor

            # Calcular dias pendentes
            dias_pendentes = 0
            if cte.envio_final:
                dias_pendentes = (datetime.now().date() - cte.envio_final).days
            elif cte.data_emissao:
                dias_pendentes = (datetime.now().date() - cte.data_emissao).days

            lista_ctes.append({
                'numero_cte': int(cte.numero_cte) if cte.numero_cte else 0,
                'data_emissao': cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                'destinatario_nome': cte.destinatario_nome or '',
                'numero_fatura': cte.numero_fatura or '',
                'valor_total': valor,
                'envio_final': cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '',
                'dias_pendentes': dias_pendentes,
                'veiculo_placa': cte.veiculo_placa or '',
                'observacao': cte.observacao or ''
            })

        return jsonify({
            'success': True,
            'dados': lista_ctes,
            'total_registros': pagination.total,
            'total_valor': total_valor,
            'pagina_atual': page,
            'total_paginas': pagination.pages,
            'tem_proxima': pagination.has_next,
            'tem_anterior': pagination.has_prev
        })

    except Exception as e:
        print(f"[ERROR] Erro na API valores pendentes: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/api/valores-pendentes/exportar/excel')
@login_required
def exportar_valores_pendentes_excel():
    """Exporta valores pendentes para Excel"""
    try:
        from io import BytesIO
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Buscar CTEs sem baixa
        cliente_filtro = request.args.get('cliente', '').strip()
        query = CTE.query.filter(CTE.data_baixa.is_(None))

        if cliente_filtro:
            query = query.filter(CTE.destinatario_nome.ilike(f'%{cliente_filtro}%'))

        ctes = query.order_by(CTE.data_emissao.asc()).all()

        if not ctes:
            return jsonify({'error': 'Nenhum valor pendente encontrado'}), 400

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valores Pendentes"

        # Estilos
        header_fill = PatternFill(start_color="0f4c75", end_color="0f4c75", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalhos
        headers = ['Nº CTE', 'Data Emissão', 'Cliente', 'Nº Fatura', 'Valor', 'Envio Final', 'Dias Pendentes', 'Veículo', 'Observação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Dados
        total_valor = 0.0
        for row_idx, cte in enumerate(ctes, 2):
            valor = float(cte.valor_total or 0)
            total_valor += valor

            dias_pendentes = 0
            if cte.envio_final:
                dias_pendentes = (datetime.now().date() - cte.envio_final).days
            elif cte.data_emissao:
                dias_pendentes = (datetime.now().date() - cte.data_emissao).days

            ws.cell(row=row_idx, column=1, value=int(cte.numero_cte) if cte.numero_cte else 0)
            ws.cell(row=row_idx, column=2, value=cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '')
            ws.cell(row=row_idx, column=3, value=cte.destinatario_nome or '')
            ws.cell(row=row_idx, column=4, value=cte.numero_fatura or '')
            ws.cell(row=row_idx, column=5, value=valor)
            ws.cell(row=row_idx, column=6, value=cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '')
            ws.cell(row=row_idx, column=7, value=dias_pendentes)
            ws.cell(row=row_idx, column=8, value=cte.veiculo_placa or '')
            ws.cell(row=row_idx, column=9, value=cte.observacao or '')

            # Formatar valor como moeda
            ws.cell(row=row_idx, column=5).number_format = 'R$ #,##0.00'

            # Aplicar bordas
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = border

        # Linha de total
        total_row = len(ctes) + 2
        ws.cell(row=total_row, column=4, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_valor).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = 'R$ #,##0.00'

        # Ajustar largura das colunas
        column_widths = [12, 15, 35, 15, 15, 15, 15, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'valores_pendentes_{timestamp}.xlsx'

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] Erro ao exportar Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@bp.route('/api/valores-pendentes/exportar/pdf')
@login_required
def exportar_valores_pendentes_pdf():
    """Exporta valores pendentes para PDF"""
    try:
        from io import BytesIO
        from flask import send_file
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        # Buscar CTEs sem baixa
        cliente_filtro = request.args.get('cliente', '').strip()
        query = CTE.query.filter(CTE.data_baixa.is_(None))

        if cliente_filtro:
            query = query.filter(CTE.destinatario_nome.ilike(f'%{cliente_filtro}%'))

        ctes = query.order_by(CTE.data_emissao.asc()).all()

        if not ctes:
            return jsonify({'error': 'Nenhum valor pendente encontrado'}), 400

        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)

        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0f4c75'),
            spaceAfter=12,
            alignment=TA_CENTER
        )

        title = Paragraph('Relatório de Valores Pendentes', title_style)
        elements.append(title)

        # Subtítulo com data
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        subtitle = Paragraph(f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}', subtitle_style)
        elements.append(subtitle)

        # Tabela de dados
        data = [['Nº CTE', 'Data', 'Cliente', 'Fatura', 'Valor', 'Envio', 'Dias', 'Veículo', 'Observação']]

        total_valor = 0.0
        for cte in ctes:
            valor = float(cte.valor_total or 0)
            total_valor += valor

            dias_pendentes = 0
            if cte.envio_final:
                dias_pendentes = (datetime.now().date() - cte.envio_final).days
            elif cte.data_emissao:
                dias_pendentes = (datetime.now().date() - cte.data_emissao).days

            data.append([
                str(int(cte.numero_cte) if cte.numero_cte else 0),
                cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                (cte.destinatario_nome or '')[:20],  # Truncar nome longo
                cte.numero_fatura or '',
                f'R$ {valor:,.2f}',
                cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '',
                str(dias_pendentes),
                cte.veiculo_placa or '',
                (cte.observacao or '')[:30]  # Truncar observação
            ])

        # Linha de total
        data.append(['', '', '', 'TOTAL:', f'R$ {total_valor:,.2f}', '', '', '', ''])

        # Criar tabela com larguras ajustadas
        table = Table(data, colWidths=[1.5*cm, 2*cm, 4.5*cm, 2*cm, 2.5*cm, 2*cm, 1.5*cm, 2*cm, 6*cm])

        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f4c75')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Dados
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Nº CTE
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Valor
            ('ALIGN', (6, 1), (6, -1), 'CENTER'),  # Dias

            # Linha de total
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4f8')),

            # Bordas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0f4c75')),

            # Alternância de cores
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f7f7f7')])
        ]))

        elements.append(table)

        # Construir PDF
        doc.build(elements)
        buffer.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'valores_pendentes_{timestamp}.pdf'

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] Erro ao exportar PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ================================
# SISTEMA DE ALERTAS - APIs Genéricas
# ================================

def _criar_exportacao_excel_alerta(ctes, titulo, filename_prefix):
    """Função genérica para criar Excel de alertas"""
    try:
        from io import BytesIO
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not ctes:
            return jsonify({'error': 'Nenhum registro encontrado'}), 400

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:31]  # Excel limita a 31 caracteres

        # Estilos
        header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Cabeçalhos
        headers = ['Nº CTE', 'Data Emissão', 'Cliente', 'Nº Fatura', 'Valor', 'Envio Final', 'Dias', 'Veículo', 'Observação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Dados
        total_valor = 0.0
        for row_idx, cte in enumerate(ctes, 2):
            valor = float(cte.valor_total or 0)
            total_valor += valor

            dias_pendentes = 0
            if cte.envio_final:
                dias_pendentes = (datetime.now().date() - cte.envio_final).days
            elif cte.data_emissao:
                dias_pendentes = (datetime.now().date() - cte.data_emissao).days

            ws.cell(row=row_idx, column=1, value=int(cte.numero_cte) if cte.numero_cte else 0)
            ws.cell(row=row_idx, column=2, value=cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '')
            ws.cell(row=row_idx, column=3, value=cte.destinatario_nome or '')
            ws.cell(row=row_idx, column=4, value=cte.numero_fatura or '')
            ws.cell(row=row_idx, column=5, value=valor)
            ws.cell(row=row_idx, column=6, value=cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '')
            ws.cell(row=row_idx, column=7, value=dias_pendentes)
            ws.cell(row=row_idx, column=8, value=cte.veiculo_placa or '')
            ws.cell(row=row_idx, column=9, value=cte.observacao or '')

            ws.cell(row=row_idx, column=5).number_format = 'R$ #,##0.00'
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = border

        # Linha de total
        total_row = len(ctes) + 2
        ws.cell(row=total_row, column=4, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_valor).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = 'R$ #,##0.00'

        # Ajustar larguras
        column_widths = [12, 15, 35, 15, 15, 15, 12, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{filename_prefix}_{timestamp}.xlsx'

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] Erro ao exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500

def _criar_exportacao_pdf_alerta(ctes, titulo, filename_prefix):
    """Função genérica para criar PDF de alertas"""
    try:
        from io import BytesIO
        from flask import send_file
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER

        if not ctes:
            return jsonify({'error': 'Nenhum registro encontrado'}), 400

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)

        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'], fontSize=16,
            textColor=colors.HexColor('#dc3545'), spaceAfter=12, alignment=TA_CENTER
        )
        title = Paragraph(titulo, title_style)
        elements.append(title)

        # Subtítulo
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Normal'], fontSize=10,
            alignment=TA_CENTER, spaceAfter=20
        )
        subtitle = Paragraph(f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}', subtitle_style)
        elements.append(subtitle)

        # Tabela
        data = [['Nº CTE', 'Data', 'Cliente', 'Fatura', 'Valor', 'Envio', 'Dias', 'Veíc.', 'Obs.']]

        total_valor = 0.0
        for cte in ctes:
            valor = float(cte.valor_total or 0)
            total_valor += valor

            dias_pendentes = 0
            if cte.envio_final:
                dias_pendentes = (datetime.now().date() - cte.envio_final).days
            elif cte.data_emissao:
                dias_pendentes = (datetime.now().date() - cte.data_emissao).days

            data.append([
                str(int(cte.numero_cte) if cte.numero_cte else 0),
                cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                (cte.destinatario_nome or '')[:18],
                cte.numero_fatura or '',
                f'R$ {valor:,.2f}',
                cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '',
                str(dias_pendentes),
                cte.veiculo_placa or '',
                (cte.observacao or '')[:25]
            ])

        data.append(['', '', '', 'TOTAL:', f'R$ {total_valor:,.2f}', '', '', '', ''])

        table = Table(data, colWidths=[1.5*cm, 2*cm, 4*cm, 2*cm, 2.5*cm, 2*cm, 1.5*cm, 1.8*cm, 5.7*cm])

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f8d7da')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f7f7f7')])
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{filename_prefix}_{timestamp}.pdf'

        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        print(f"[ERROR] Erro ao exportar PDF: {e}")
        return jsonify({'error': str(e)}), 500

# ================================
# 1º ENVIO PENDENTE
# ================================

@bp.route('/api/primeiro-envio-pendente')
@login_required
def api_primeiro_envio_pendente():
    """Lista CTEs com 1º envio pendente"""
    try:
        hoje = datetime.now().date()
        data_limite = hoje - timedelta(days=10)

        query = CTE.query.filter(
            CTE.data_emissao < data_limite,
            CTE.primeiro_envio.is_(None)
        ).order_by(CTE.data_emissao.asc())

        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        lista = []
        total_valor = 0.0
        for cte in pagination.items:
            valor = float(cte.valor_total or 0)
            total_valor += valor
            dias = (hoje - cte.data_emissao).days if cte.data_emissao else 0

            lista.append({
                'numero_cte': int(cte.numero_cte) if cte.numero_cte else 0,
                'data_emissao': cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                'destinatario_nome': cte.destinatario_nome or '',
                'numero_fatura': cte.numero_fatura or '',
                'valor_total': valor,
                'envio_final': cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '',
                'dias_pendentes': dias,
                'veiculo_placa': cte.veiculo_placa or '',
                'observacao': cte.observacao or ''
            })

        return jsonify({
            'success': True,
            'dados': lista,
            'total_registros': pagination.total,
            'total_valor': total_valor,
            'pagina_atual': page,
            'total_paginas': pagination.pages
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/primeiro-envio-pendente/exportar/excel')
@login_required
def exportar_primeiro_envio_excel():
    """Exporta 1º envio pendente para Excel"""
    hoje = datetime.now().date()
    data_limite = hoje - timedelta(days=10)
    ctes = CTE.query.filter(CTE.data_emissao < data_limite, CTE.primeiro_envio.is_(None)).all()
    return _criar_exportacao_excel_alerta(ctes, '1º Envio Pendente', 'primeiro_envio_pendente')

@bp.route('/api/primeiro-envio-pendente/exportar/pdf')
@login_required
def exportar_primeiro_envio_pdf():
    """Exporta 1º envio pendente para PDF"""
    hoje = datetime.now().date()
    data_limite = hoje - timedelta(days=10)
    ctes = CTE.query.filter(CTE.data_emissao < data_limite, CTE.primeiro_envio.is_(None)).all()
    return _criar_exportacao_pdf_alerta(ctes, 'Relatório: 1º Envio Pendente', 'primeiro_envio_pendente')

# ================================
# ENVIO FINAL PENDENTE
# ================================

@bp.route('/api/envio-final-pendente')
@login_required
def api_envio_final_pendente():
    """Lista CTEs com envio final pendente (todos sem envio final)"""
    try:
        hoje = datetime.now().date()

        # Buscar TODOS os CTEs sem envio final (sem filtro de data)
        query = CTE.query.filter(
            CTE.envio_final.is_(None)
        ).order_by(CTE.data_emissao.asc())

        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        lista = []
        total_valor = 0.0
        for cte in pagination.items:
            valor = float(cte.valor_total or 0)
            total_valor += valor
            # Calcular dias desde a emissão
            dias = (hoje - cte.data_emissao).days if cte.data_emissao else 0

            lista.append({
                'numero_cte': int(cte.numero_cte) if cte.numero_cte else 0,
                'data_emissao': cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                'destinatario_nome': cte.destinatario_nome or '',
                'numero_fatura': cte.numero_fatura or '',
                'valor_total': valor,
                'envio_final': '',
                'dias_pendentes': dias,
                'veiculo_placa': cte.veiculo_placa or '',
                'observacao': cte.observacao or ''
            })

        return jsonify({
            'success': True,
            'dados': lista,
            'total_registros': pagination.total,
            'total_valor': total_valor,
            'pagina_atual': page,
            'total_paginas': pagination.pages
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/envio-final-pendente/exportar/excel')
@login_required
def exportar_envio_final_excel():
    """Exporta TODOS os CTEs sem envio final para Excel"""
    # Buscar TODOS os CTEs sem envio final (sem filtro de data)
    ctes = CTE.query.filter(CTE.envio_final.is_(None)).order_by(CTE.data_emissao.asc()).all()
    return _criar_exportacao_excel_alerta(ctes, 'Envio Final Pendente', 'envio_final_pendente')

@bp.route('/api/envio-final-pendente/exportar/pdf')
@login_required
def exportar_envio_final_pdf():
    """Exporta TODOS os CTEs sem envio final para PDF"""
    # Buscar TODOS os CTEs sem envio final (sem filtro de data)
    ctes = CTE.query.filter(CTE.envio_final.is_(None)).order_by(CTE.data_emissao.asc()).all()
    return _criar_exportacao_pdf_alerta(ctes, 'Relatório: Envio Final Pendente', 'envio_final_pendente')

# ================================
# FATURAS VENCIDAS
# ================================

@bp.route('/api/faturas-vencidas')
@login_required
def api_faturas_vencidas():
    """Lista faturas vencidas (90+ dias)"""
    try:
        hoje = datetime.now().date()
        data_limite = hoje - timedelta(days=90)

        query = CTE.query.filter(
            CTE.envio_final < data_limite,
            CTE.data_baixa.is_(None)
        ).order_by(CTE.envio_final.asc())

        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        lista = []
        total_valor = 0.0
        for cte in pagination.items:
            valor = float(cte.valor_total or 0)
            total_valor += valor
            dias = (hoje - cte.envio_final).days if cte.envio_final else 0

            lista.append({
                'numero_cte': int(cte.numero_cte) if cte.numero_cte else 0,
                'data_emissao': cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                'destinatario_nome': cte.destinatario_nome or '',
                'numero_fatura': cte.numero_fatura or '',
                'valor_total': valor,
                'envio_final': cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '',
                'dias_pendentes': dias,
                'veiculo_placa': cte.veiculo_placa or '',
                'observacao': cte.observacao or ''
            })

        return jsonify({
            'success': True,
            'dados': lista,
            'total_registros': pagination.total,
            'total_valor': total_valor,
            'pagina_atual': page,
            'total_paginas': pagination.pages
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/faturas-vencidas/exportar/excel')
@login_required
def exportar_faturas_vencidas_excel():
    """Exporta faturas vencidas para Excel"""
    hoje = datetime.now().date()
    data_limite = hoje - timedelta(days=90)
    ctes = CTE.query.filter(CTE.envio_final < data_limite, CTE.data_baixa.is_(None)).all()
    return _criar_exportacao_excel_alerta(ctes, 'Faturas Vencidas', 'faturas_vencidas')

@bp.route('/api/faturas-vencidas/exportar/pdf')
@login_required
def exportar_faturas_vencidas_pdf():
    """Exporta faturas vencidas para PDF"""
    hoje = datetime.now().date()
    data_limite = hoje - timedelta(days=90)
    ctes = CTE.query.filter(CTE.envio_final < data_limite, CTE.data_baixa.is_(None)).all()
    return _criar_exportacao_pdf_alerta(ctes, 'Relatório: Faturas Vencidas (90+ dias)', 'faturas_vencidas')

# ================================
# CTES SEM FATURAS
# ================================

@bp.route('/api/ctes-sem-faturas')
@login_required
def api_ctes_sem_faturas():
    """Lista CTEs sem número de fatura"""
    try:
        hoje = datetime.now().date()
        data_limite = hoje - timedelta(days=3)

        query = CTE.query.filter(
            CTE.data_atesto < data_limite,
            db.or_(CTE.numero_fatura.is_(None), CTE.numero_fatura == '')
        ).order_by(CTE.data_atesto.asc())

        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        lista = []
        total_valor = 0.0
        for cte in pagination.items:
            valor = float(cte.valor_total or 0)
            total_valor += valor
            dias = (hoje - cte.data_atesto).days if cte.data_atesto else 0

            lista.append({
                'numero_cte': int(cte.numero_cte) if cte.numero_cte else 0,
                'data_emissao': cte.data_emissao.strftime('%d/%m/%Y') if cte.data_emissao else '',
                'destinatario_nome': cte.destinatario_nome or '',
                'numero_fatura': '',
                'valor_total': valor,
                'envio_final': cte.envio_final.strftime('%d/%m/%Y') if cte.envio_final else '',
                'dias_pendentes': dias,
                'veiculo_placa': cte.veiculo_placa or '',
                'observacao': cte.observacao or ''
            })

        return jsonify({
            'success': True,
            'dados': lista,
            'total_registros': pagination.total,
            'total_valor': total_valor,
            'pagina_atual': page,
            'total_paginas': pagination.pages
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/ctes-sem-faturas/exportar/excel')
@login_required
def exportar_ctes_sem_faturas_excel():
    """Exporta CTEs sem faturas para Excel"""
    hoje = datetime.now().date()
    data_limite = hoje - timedelta(days=3)
    ctes = CTE.query.filter(
        CTE.data_atesto < data_limite,
        db.or_(CTE.numero_fatura.is_(None), CTE.numero_fatura == '')
    ).all()
    return _criar_exportacao_excel_alerta(ctes, 'CTEs sem Faturas', 'ctes_sem_faturas')

@bp.route('/api/ctes-sem-faturas/exportar/pdf')
@login_required
def exportar_ctes_sem_faturas_pdf():
    """Exporta CTEs sem faturas para PDF"""
    hoje = datetime.now().date()
    data_limite = hoje - timedelta(days=3)
    ctes = CTE.query.filter(
        CTE.data_atesto < data_limite,
        db.or_(CTE.numero_fatura.is_(None), CTE.numero_fatura == '')
    ).all()
    return _criar_exportacao_pdf_alerta(ctes, 'Relatório: CTEs sem Faturas', 'ctes_sem_faturas')

# ================================
# API DE RESUMO DOS ALERTAS (TOTAIS REAIS)
# ================================

@bp.route('/api/alertas/resumo')
@login_required
def api_alertas_resumo():
    """Retorna totais reais de cada tipo de alerta para os cards"""
    try:
        hoje = datetime.now().date()

        # 1º Envio Pendente (emitidos há mais de 10 dias sem primeiro envio)
        data_limite_primeiro = hoje - timedelta(days=10)
        query_primeiro = CTE.query.filter(
            CTE.data_emissao < data_limite_primeiro,
            CTE.primeiro_envio.is_(None)
        )
        qtd_primeiro = query_primeiro.count()
        valor_primeiro = db.session.query(db.func.sum(CTE.valor_total)).filter(
            CTE.data_emissao < data_limite_primeiro,
            CTE.primeiro_envio.is_(None)
        ).scalar() or 0.0

        # Envio Final Pendente (TODOS os CTEs sem envio final)
        # IMPORTANTE: Usar mesmos critérios das APIs de listagem/exportação
        query_final = CTE.query.filter(
            CTE.envio_final.is_(None)
        )
        qtd_final = query_final.count()
        valor_final = db.session.query(db.func.sum(CTE.valor_total)).filter(
            CTE.envio_final.is_(None)
        ).scalar() or 0.0

        # Faturas Vencidas (envio final há mais de 90 dias sem data de baixa)
        # IMPORTANTE: Usar mesmos critérios das APIs de listagem/exportação
        data_limite_vencidas = hoje - timedelta(days=90)
        query_vencidas = CTE.query.filter(
            CTE.envio_final < data_limite_vencidas,
            CTE.data_baixa.is_(None)
        )
        qtd_vencidas = query_vencidas.count()
        valor_vencidas = db.session.query(db.func.sum(CTE.valor_total)).filter(
            CTE.envio_final < data_limite_vencidas,
            CTE.data_baixa.is_(None)
        ).scalar() or 0.0

        # CTEs sem Faturas (atesto há mais de 3 dias sem número de fatura)
        data_limite_sem_fatura = hoje - timedelta(days=3)
        query_sem_fatura = CTE.query.filter(
            CTE.data_atesto < data_limite_sem_fatura,
            db.or_(CTE.numero_fatura.is_(None), CTE.numero_fatura == '')
        )
        qtd_sem_fatura = query_sem_fatura.count()
        valor_sem_fatura = db.session.query(db.func.sum(CTE.valor_total)).filter(
            CTE.data_atesto < data_limite_sem_fatura,
            db.or_(CTE.numero_fatura.is_(None), CTE.numero_fatura == '')
        ).scalar() or 0.0

        return jsonify({
            'success': True,
            'primeiro_envio_pendente': {
                'quantidade': qtd_primeiro,
                'valor_total': float(valor_primeiro)
            },
            'envio_final_pendente': {
                'quantidade': qtd_final,
                'valor_total': float(valor_final)
            },
            'faturas_vencidas': {
                'quantidade': qtd_vencidas,
                'valor_total': float(valor_vencidas)
            },
            'ctes_sem_faturas': {
                'quantidade': qtd_sem_fatura,
                'valor_total': float(valor_sem_fatura)
            }
        })
    except Exception as e:
        print(f"[ERROR] Erro ao calcular resumo de alertas: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

    # ================================
# ADICIONAR ESTAS FUNÇÕES no arquivo dashboard.py
# ================================

def calcular_alertas_inteligentes(df):
    """Sistema de alertas inteligentes - VERSÃO CORRIGIDA"""
    print(f"[ALERT] Iniciando cálculo de alertas para {len(df)} registros")
    
    alertas = {
        'primeiro_envio_pendente': {'qtd': 0, 'valor': 0.0, 'lista': []},
        'envio_final_pendente': {'qtd': 0, 'valor': 0.0, 'lista': []},
        'faturas_vencidas': {'qtd': 0, 'valor': 0.0, 'lista': []},
        'ctes_sem_faturas': {'qtd': 0, 'valor': 0.0, 'lista': []}
    }

    if df.empty:
        print("[WARN] DataFrame vazio para alertas")
        return alertas

    hoje = pd.Timestamp.now().normalize()

    try:
        # 1. Primeiro envio pendente (10 dias após emissão) - ✅ MANTIDO
        print("[SEARCH] Calculando primeiro envio pendente...")
        mask_primeiro_envio = (
            df['data_emissao'].notna() & 
            ((hoje - df['data_emissao']).dt.days > 10) &
            df['primeiro_envio'].isna()
        )
        if mask_primeiro_envio.any():
            ctes_problema = df[mask_primeiro_envio]
            lista_segura = []
            for _, row in ctes_problema.head(10).iterrows():  # Limitar a 10 por segurança
                try:
                    item = {
                        'numero_cte': int(row['numero_cte']),
                        'destinatario_nome': str(row['destinatario_nome']),
                        'valor_total': float(row['valor_total']),
                        'data_emissao': row['data_emissao'].isoformat() if pd.notna(row['data_emissao']) else None
                    }
                    lista_segura.append(item)
                except Exception as e:
                    print(f"[WARN] Erro ao processar CTE: {e}")
                    continue

            alertas['primeiro_envio_pendente'] = {
                'qtd': len(ctes_problema),
                'valor': float(ctes_problema['valor_total'].sum()),
                'lista': lista_segura
            }
            print(f"[OK] Primeiro envio pendente: {len(ctes_problema)} CTEs")

        # 2. Envio Final Pendente - 🔧 CORRIGIDO: 1 dia após atesto (era 5)
        print("[SEARCH] Calculando envio final pendente...")
        mask_envio_final = (
            df['data_atesto'].notna() & 
            ((hoje - df['data_atesto']).dt.days > 1) &  # MUDANÇA: 5 → 1
            df['envio_final'].isna()
        )
        if mask_envio_final.any():
            ctes_problema = df[mask_envio_final]
            lista_segura = []
            for _, row in ctes_problema.head(10).iterrows():
                try:
                    item = {
                        'numero_cte': int(row['numero_cte']),
                        'destinatario_nome': str(row['destinatario_nome']),
                        'valor_total': float(row['valor_total']),
                        'data_atesto': row['data_atesto'].isoformat() if pd.notna(row['data_atesto']) else None
                    }
                    lista_segura.append(item)
                except Exception as e:
                    print(f"[WARN] Erro ao processar CTE: {e}")
                    continue

            alertas['envio_final_pendente'] = {
                'qtd': len(ctes_problema),
                'valor': float(ctes_problema['valor_total'].sum()),
                'lista': lista_segura
            }
            print(f"[OK] Envio final pendente: {len(ctes_problema)} CTEs")

        # 3. Faturas vencidas - 🔧 CORRIGIDO: 90 dias após ENVIO FINAL (era após atesto)
        print("[SEARCH] Calculando faturas vencidas...")
        mask_vencidas = (
            df['envio_final'].notna() &  # MUDANÇA: data_atesto → envio_final
            ((hoje - df['envio_final']).dt.days > 90) &  # MUDANÇA: data_atesto → envio_final
            df['data_baixa'].isna()
        )
        if mask_vencidas.any():
            ctes_problema = df[mask_vencidas]
            lista_segura = []
            for _, row in ctes_problema.head(10).iterrows():
                try:
                    item = {
                        'numero_cte': int(row['numero_cte']),
                        'destinatario_nome': str(row['destinatario_nome']),
                        'valor_total': float(row['valor_total']),
                        'envio_final': row['envio_final'].isoformat() if pd.notna(row['envio_final']) else None  # MUDANÇA: data_atesto → envio_final
                    }
                    lista_segura.append(item)
                except Exception as e:
                    print(f"[WARN] Erro ao processar CTE: {e}")
                    continue

            alertas['faturas_vencidas'] = {
                'qtd': len(ctes_problema),
                'valor': float(ctes_problema['valor_total'].sum()),
                'lista': lista_segura
            }
            print(f"[OK] Faturas vencidas: {len(ctes_problema)} CTEs")

        # 4. CTEs sem faturas (3 dias após atesto) - ✅ MANTIDO
        print("[SEARCH] Calculando CTEs sem faturas...")
        mask_sem_faturas = (
            df['data_atesto'].notna() & 
            ((hoje - df['data_atesto']).dt.days > 3) &
            (df['numero_fatura'].isna() | (df['numero_fatura'] == ''))
        )
        if mask_sem_faturas.any():
            ctes_problema = df[mask_sem_faturas]
            lista_segura = []
            for _, row in ctes_problema.head(10).iterrows():
                try:
                    item = {
                        'numero_cte': int(row['numero_cte']),
                        'destinatario_nome': str(row['destinatario_nome']),
                        'valor_total': float(row['valor_total']),
                        'data_atesto': row['data_atesto'].isoformat() if pd.notna(row['data_atesto']) else None
                    }
                    lista_segura.append(item)
                except Exception as e:
                    print(f"[WARN] Erro ao processar CTE: {e}")
                    continue

            alertas['ctes_sem_faturas'] = {
                'qtd': len(ctes_problema),
                'valor': float(ctes_problema['valor_total'].sum()),
                'lista': lista_segura
            }
            print(f"[OK] CTEs sem faturas: {len(ctes_problema)} CTEs")

        print("[OK] Todos os alertas calculados com sucesso!")

    except Exception as e:
        print(f"[WARN] Erro no cálculo de alertas: {str(e)}")
        import traceback
        traceback.print_exc()

    return alertas


def _variacoes(df):
    """Calcula variações de tempo entre processos"""
    print(f"[TIME] Calculando variações para {len(df)} registros")
    
    configs = [
        ('rq_tmc_primeiro_envio', 'RQ/TMC - 1º Envio', 'data_rq_tmc', 'primeiro_envio', 3),
        ('primeiro_envio_atesto', '1º Envio - Atesto', 'primeiro_envio', 'data_atesto', 7),
        ('atesto_envio_final', 'Atesto - Envio Final', 'data_atesto', 'envio_final', 2),
        ('cte_inclusao_fatura', 'CTE - Inclusão Fatura', 'data_emissao', 'data_inclusao_fatura', 2),
        ('cte_baixa', 'CTE - Baixa', 'data_emissao', 'data_baixa', 30),
    ]
    
    out = {}
    if df.empty:
        return out

    try:
        for code, nome, inicio, fim, meta in configs:
            if inicio in df.columns and fim in df.columns:
                m = df[inicio].notna() & df[fim].notna()
                if not m.any():
                    continue
                
                _dif = (df.loc[m, fim] - df.loc[m, inicio]).dt.days
                _dif = _dif[_dif >= 0]  # Apenas diferenças positivas
                
                if len(_dif) == 0:
                    continue
                
                media = float(_dif.mean())
                mediana = float(_dif.median())
                
                # Determinar performance
                if media <= meta:
                    perf = 'excelente'
                elif media <= meta * 1.5:
                    perf = 'bom'
                elif media <= meta * 2:
                    perf = 'atencao'
                else:
                    perf = 'critico'
                
                out[code] = {
                    'nome': nome,
                    'media': round(media, 1),
                    'mediana': round(mediana, 1),
                    'qtd': int(len(_dif)),
                    'meta_dias': int(meta),
                    'performance': perf,
                    'min': int(_dif.min()),
                    'max': int(_dif.max())
                }
                print(f"[OK] {nome}: {len(_dif)} registros, média {media:.1f} dias")
                
    except Exception as e:
        print(f"[WARN] Erro no cálculo de variações: {e}")

    return out


def _graficos(df):
    """Gera dados para gráficos do dashboard"""
    print(f"[GRAPH] Gerando gráficos para {len(df)} registros")
    
    graficos = {
        'evolucao_mensal': {'labels': [], 'valores': [], 'quantidades': []},
        'top_clientes': {'labels': [], 'valores': []},
        'distribuicao_status': {
            'baixas': {'labels': [], 'valores': []},
            'processos': {'labels': [], 'valores': []}
        },
        'performance_veiculos': {'labels': [], 'valores': [], 'quantidades': []}
    }
    
    if df.empty:
        return graficos

    try:
        valores = pd.to_numeric(df['valor_total'], errors='coerce').fillna(0)

        # Evolução mensal
        if 'data_emissao' in df.columns and df['data_emissao'].notna().any():
            try:
                tmp = df[df['data_emissao'].notna() & valores.notna()].copy()
                if not tmp.empty:
                    tmp['mes_ano'] = tmp['data_emissao'].dt.to_period('M')
                    receita = tmp.groupby('mes_ano').agg(
                        valor_total=('valor_total', lambda x: pd.to_numeric(x, errors='coerce').sum()),
                        qtd=('numero_cte', 'count')
                    ).reset_index().tail(12)
                    
                    if not receita.empty:
                        graficos['evolucao_mensal'] = {
                            'labels': [str(p) for p in receita['mes_ano']],
                            'valores': [float(v) for v in receita['valor_total']],
                            'quantidades': [int(q) for q in receita['qtd']]
                        }
                        print(f"[OK] Evolução mensal: {len(receita)} períodos")
            except Exception as e:
                print(f"[WARN] Erro na evolução mensal: {e}")

        # Top clientes
        if 'destinatario_nome' in df.columns:
            try:
                tmp_df = df.copy()
                tmp_df['valor_total'] = valores
                top = tmp_df.groupby('destinatario_nome')['valor_total'].sum().sort_values(ascending=False).head(5)
                if not top.empty:
                    graficos['top_clientes'] = {
                        'labels': [str(i) for i in top.index],
                        'valores': [float(v) for v in top.values]
                    }
                    print(f"[OK] Top clientes: {len(top)} clientes")
            except Exception as e:
                print(f"[WARN] Erro no top clientes: {e}")

        # Distribuição de status
        try:
            com_baixa = int(df['data_baixa'].notna().sum()) if 'data_baixa' in df.columns else 0
            sem_baixa = int(len(df) - com_baixa)
            
            proc_compl = 0
            if all(col in df.columns for col in ['data_emissao', 'primeiro_envio', 'data_atesto', 'envio_final']):
                proc_compl = int((df['data_emissao'].notna() & df['primeiro_envio'].notna() &
                                  df['data_atesto'].notna() & df['envio_final'].notna()).sum())
            proc_incompl = int(len(df) - proc_compl)
            
            graficos['distribuicao_status'] = {
                'baixas': {'labels': ['Com Baixa', 'Sem Baixa'], 'valores': [com_baixa, sem_baixa]},
                'processos': {'labels': ['Completos', 'Incompletos'], 'valores': [proc_compl, proc_incompl]}
            }
            print(f"[OK] Distribuição: {com_baixa} com baixa, {proc_compl} completos")
        except Exception as e:
            print(f"[WARN] Erro na distribuição de status: {e}")

        # Performance de veículos
        if 'veiculo_placa' in df.columns:
            try:
                tmp_df = df.copy()
                tmp_df['valor_total'] = valores
                v = tmp_df.groupby('veiculo_placa').agg(
                    valor_total=('valor_total', 'sum'),
                    qtd=('numero_cte', 'count')
                ).sort_values('valor_total', ascending=False).head(10)
                
                if not v.empty:
                    graficos['performance_veiculos'] = {
                        'labels': [str(i) for i in v.index],
                        'valores': [float(x) for x in v['valor_total']],
                        'quantidades': [int(x) for x in v['qtd']]
                    }
                    print(f"[OK] Performance veículos: {len(v)} veículos")
            except Exception as e:
                print(f"[WARN] Erro na performance de veículos: {e}")

        print("[OK] Gráficos gerados com sucesso!")

    except Exception as e:
        print(f"[WARN] Erro geral nos gráficos: {e}")

    return graficos