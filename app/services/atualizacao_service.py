# app/services/atualizacao_service.py
from __future__ import annotations

print("DEBUG: Starting imports...")
from io import BytesIO
from typing import Dict, List, Tuple, Optional
from datetime import datetime
print("DEBUG: Basic imports done")

import pandas as pd
print("DEBUG: pandas imported")

try:
    from app import db
    print("DEBUG: app.db imported")
    from app.models.cte import CTE
    print("DEBUG: CTE imported")
except ImportError as e:
    print(f"Import error: {e}")
    raise

print("DEBUG: Defining constants...")
ALIAS_COLUNAS = {
    # Cabeçalhos "humanos" -> campos do modelo
    "Número CTE": "numero_cte",
    "Cliente": "destinatario_nome",
    "Destinatário": "destinatario_nome",
    "Placa Veículo": "veiculo_placa",
    "Placa": "veiculo_placa", 
    "Valor Total": "valor_total",
    "Valor": "valor_total",
    "Data Emissão": "data_emissao",
    "Data de Emissão": "data_emissao",
    "Data Baixa": "data_baixa",
    "Data de Baixa": "data_baixa",
    "Número Fatura": "numero_fatura",
    "Fatura": "numero_fatura",
    "Data Inclusão Fatura": "data_inclusao_fatura",
    "Data Inclusão na Fatura": "data_inclusao_fatura",
    "Data Envio Processo": "data_envio_processo",
    "Data de Envio do Processo": "data_envio_processo",
    "Primeiro Envio": "primeiro_envio",
    "Data Primeiro Envio": "primeiro_envio",
    "Data RQ/TMC": "data_rq_tmc",
    "Data RQ TMC": "data_rq_tmc",
    "RQ/TMC": "data_rq_tmc",
    "Data Atesto": "data_atesto",
    "Data de Atesto": "data_atesto",
    "Atesto": "data_atesto",
    "Envio Final": "envio_final",
    "Data Envio Final": "envio_final",
    "Observação": "observacao",
    "Observações": "observacao",
    "Obs": "observacao",
    "Origem dos Dados": "origem_dados",
    "Origem": "origem_dados",
}

COLUNAS_COMPLETAS = [
    "numero_cte", "destinatario_nome", "veiculo_placa", "valor_total",
    "data_emissao", "data_baixa", 
    "numero_fatura", "data_inclusao_fatura", "data_envio_processo",
    "primeiro_envio", "data_rq_tmc", "data_atesto", "envio_final",
    "observacao", "origem_dados",
]


class AtualizacaoService:
    """Importação/atualização em lote de CTEs (CSV + Excel)."""

    print("DEBUG: Inside class definition...")

    @staticmethod
    def _read_as_dataframe(file_storage) -> Tuple[bool, str, Optional[pd.DataFrame]]:
        """Lê CSV/Excel para DataFrame."""
        if not file_storage or not getattr(file_storage, "filename", ""):
            return False, "Nenhum arquivo enviado", None

        nome = file_storage.filename.lower()

        try:
            if nome.endswith(".csv"):
                raw = file_storage.read()
                if hasattr(file_storage, 'stream'):
                    file_storage.stream.seek(0)
                
                raw_str = raw.decode('utf-8-sig')
                
                if ';' in raw_str and raw_str.count(';') > raw_str.count(','):
                    sep = ';'
                else:
                    sep = ','
                
                try:
                    df = pd.read_csv(BytesIO(raw), sep=sep, encoding="utf-8-sig")
                    if len(df.columns) == 1 and sep == ';':
                        df = pd.read_csv(BytesIO(raw), sep=",", encoding="utf-8-sig")
                    return True, f"CSV lido (separador: '{sep}')", df
                except Exception as e:
                    try:
                        alt_sep = ',' if sep == ';' else ';'
                        df = pd.read_csv(BytesIO(raw), sep=alt_sep, encoding="utf-8-sig")
                        return True, f"CSV lido (separador alternativo: '{alt_sep}')", df
                    except Exception:
                        return False, f"Falha ao ler CSV: {e}", None

            elif nome.endswith(".xlsx") or nome.endswith(".xls"):
                df = pd.read_excel(file_storage, sheet_name=0, engine=None)
                return True, "Excel lido", df

            else:
                return False, "Formato não suportado. Envie CSV ou Excel (.xlsx/.xls).", None

        except Exception as e:
            return False, f"Falha ao ler o arquivo: {e}", None

    @staticmethod
    def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Normaliza cabeçalhos para os campos reais do modelo."""
        if df is None or df.empty:
            return df

        mapping = {}
        for col in df.columns:
            mapping[col] = ALIAS_COLUNAS.get(col, col)

        df = df.rename(columns=mapping)
        return df

    @staticmethod
    def validar_arquivo(file_storage) -> Tuple[bool, str, Dict]:
        """Valida rapidamente o arquivo e dá estatísticas."""
        if hasattr(file_storage, 'seek'):
            file_storage.seek(0)
            
        ok, msg, df = AtualizacaoService._read_as_dataframe(file_storage)
        if not ok:
            return False, msg, {}

        df = AtualizacaoService._normalize_columns(df)

        colunas_originais = list(df.columns) if df is not None else []
        
        if df is None or df.empty:
            return False, "Arquivo vazio ou inválido", {}

        if "numero_cte" not in df.columns:
            return False, f"Arquivo sem coluna 'numero_cte' (ou 'Número CTE'). Colunas encontradas: {colunas_originais}", {}

        total = len(df)
        preview = df.head(5).to_dict("records")

        return True, "Arquivo válido", {
            "linhas": total,
            "colunas": list(df.columns),
            "preview": preview,
        }

    @staticmethod
    def processar_atualizacao(file_storage, modo: str = "alterar") -> Dict:
        """Processa o arquivo: altera registros existentes e opcionalmente insere novos."""
        resultado = {
            "sucesso": False,
            "processados": 0,
            "atualizados": 0,
            "inseridos": 0,
            "ignorados": 0,
            "erros": 0,
            "detalhes": [],
        }

        if hasattr(file_storage, 'seek'):
            file_storage.seek(0)

        ok, msg, df = AtualizacaoService._read_as_dataframe(file_storage)
        if not ok:
            resultado["detalhes"].append({"erro": msg})
            return resultado

        df = AtualizacaoService._normalize_columns(df)

        if df.empty:
            resultado["detalhes"].append({"erro": "Nenhum registro no arquivo"})
            return resultado

        for col in COLUNAS_COMPLETAS:
            if col not in df.columns:
                df[col] = None

        try:
            for _, row in df.iterrows():
                resultado["processados"] += 1

                # Extrai dados brutos
                dados_brutos = {k: row.get(k) for k in COLUNAS_COMPLETAS if k in df.columns}
                
                # Normaliza os dados usando as funções de limpeza
                dados = AtualizacaoService._normalizar_dados_linha(dados_brutos)
                numero = dados.get("numero_cte")

                if numero is None:
                    resultado["ignorados"] += 1
                    resultado["detalhes"].append({
                        "cte": None, "sucesso": False,
                        "mensagem": "Linha sem número do CTE válido"
                    })
                    continue

                # O número já foi limpo pela função de normalização
                if not isinstance(numero, int):
                    resultado["erros"] += 1
                    resultado["detalhes"].append({
                        "cte": str(numero), "sucesso": False,
                        "mensagem": "Número do CTE inválido após normalização"
                    })
                    continue

                cte = CTE.buscar_por_numero(numero)

                if cte:
                    ok_upd, msg_upd = cte.atualizar(dados)
                    if ok_upd:
                        resultado["atualizados"] += 1
                        resultado["detalhes"].append({
                            "cte": numero, "sucesso": True, "mensagem": "Atualizado"
                        })
                    else:
                        resultado["erros"] += 1
                        resultado["detalhes"].append({
                            "cte": numero, "sucesso": False, "mensagem": msg_upd
                        })
                else:
                    if modo.lower() in ("upsert", "inserir", "criar"):
                        ok_ins, obj_ou_msg = CTE.criar_cte(dados)
                        if ok_ins:
                            resultado["inseridos"] += 1
                            resultado["detalhes"].append({
                                "cte": numero, "sucesso": True, "mensagem": "Criado"
                            })
                        else:
                            resultado["erros"] += 1
                            resultado["detalhes"].append({
                                "cte": numero, "sucesso": False, "mensagem": obj_ou_msg
                            })
                    else:
                        resultado["ignorados"] += 1
                        resultado["detalhes"].append({
                            "cte": numero, "sucesso": False,
                            "mensagem": "CTE não existe (modo apenas atualizar)"
                        })

            db.session.commit()
            resultado["sucesso"] = True

        except Exception as e:
            db.session.rollback()
            resultado["detalhes"].append({"erro": f"Erro no processamento: {e}"})

        return resultado

    @staticmethod
    def template_csv() -> str:
        """Gera template CSV completo para download."""
        headers = [
            "Número CTE",
            "Destinatário", 
            "Placa Veículo",
            "Valor Total",
            "Data Emissão",
            "Data Baixa",
            "Número Fatura",
            "Data Inclusão Fatura",
            "Data Envio Processo", 
            "Primeiro Envio",
            "Data RQ/TMC",
            "Data Atesto",
            "Envio Final",
            "Observação",
            "Origem dos Dados"
        ]
        
        # Linha de cabeçalho
        csv_content = ";".join(headers) + "\n"
        
        # Comentários explicativos sobre formato esperado
        csv_content += ";".join([
            "# FORMATOS ESPERADOS:",
            "Razão Social Completa",
            "ABC-1234",
            "1500.50 (decimal com ponto)",
            "AAAA-MM-DD (2025-01-15)",
            "AAAA-MM-DD ou vazio",
            "Número/Código da Fatura",
            "AAAA-MM-DD ou vazio",
            "AAAA-MM-DD ou vazio",
            "AAAA-MM-DD ou vazio",
            "AAAA-MM-DD ou vazio",
            "AAAA-MM-DD ou vazio",
            "AAAA-MM-DD ou vazio",
            "Texto livre",
            "Sistema/CSV/Manual"
        ]) + "\n"
        
        # Linha de exemplo com dados de amostra corretamente formatados
        exemplo = [
            "12345",
            "Cliente Exemplo Ltda",
            "ABC-1234",
            "1500.50",
            "2025-01-15",
            "2025-01-20", 
            "FAT001",
            "2025-01-22",
            "2025-01-25",
            "2025-01-26",
            "2025-01-28",
            "2025-01-30",
            "2025-02-01",
            "Exemplo de observação do CTE",
            "Sistema"
        ]
        csv_content += ";".join(exemplo) + "\n"
        
        # Exemplo adicional com campos opcionais vazios
        exemplo2 = [
            "12346",
            "Outro Cliente SA",
            "XYZ-5678",
            "2300.75",
            "2025-01-16",
            "",  # Data Baixa vazia
            "",  # Número Fatura vazio
            "",  # Data Inclusão Fatura vazia
            "",  # Data Envio Processo vazia
            "",  # Primeiro Envio vazio
            "",  # Data RQ/TMC vazia
            "",  # Data Atesto vazia
            "",  # Envio Final vazio
            "CTE ainda em processo",
            "CSV"
        ]
        csv_content += ";".join(exemplo2) + "\n"
        
        return csv_content

    @staticmethod
    def template_excel() -> BytesIO:
        """Gera template Excel completo para download."""
        headers = [
            "Número CTE",
            "Destinatário", 
            "Placa Veículo",
            "Valor Total",
            "Data Emissão",
            "Data Baixa",
            "Número Fatura",
            "Data Inclusão Fatura",
            "Data Envio Processo", 
            "Primeiro Envio",
            "Data RQ/TMC",
            "Data Atesto",
            "Envio Final",
            "Observação",
            "Origem dos Dados"
        ]
        
        # Linha de instruções sobre formatos
        instrucoes = {
            "Número CTE": "Número inteiro",
            "Destinatário": "Razão Social Completa",
            "Placa Veículo": "ABC-1234",
            "Valor Total": "1500.50 (decimal)",
            "Data Emissão": "AAAA-MM-DD",
            "Data Baixa": "AAAA-MM-DD ou vazio",
            "Número Fatura": "Código/Número",
            "Data Inclusão Fatura": "AAAA-MM-DD ou vazio",
            "Data Envio Processo": "AAAA-MM-DD ou vazio",
            "Primeiro Envio": "AAAA-MM-DD ou vazio",
            "Data RQ/TMC": "AAAA-MM-DD ou vazio",
            "Data Atesto": "AAAA-MM-DD ou vazio",
            "Envio Final": "AAAA-MM-DD ou vazio",
            "Observação": "Texto livre",
            "Origem dos Dados": "Sistema/CSV/Manual"
        }
        
        # Dados de exemplo
        dados_exemplo = [
            instrucoes,  # Primeira linha com instruções
            {
                "Número CTE": 12345,
                "Destinatário": "Cliente Exemplo Ltda",
                "Placa Veículo": "ABC-1234",
                "Valor Total": 1500.50,
                "Data Emissão": "2025-01-15",
                "Data Baixa": "2025-01-20",
                "Número Fatura": "FAT001",
                "Data Inclusão Fatura": "2025-01-22",
                "Data Envio Processo": "2025-01-25",
                "Primeiro Envio": "2025-01-26",
                "Data RQ/TMC": "2025-01-28",
                "Data Atesto": "2025-01-30",
                "Envio Final": "2025-02-01",
                "Observação": "Exemplo de observação do CTE",
                "Origem dos Dados": "Sistema"
            },
            {
                "Número CTE": 12346,
                "Destinatário": "Outro Cliente SA",
                "Placa Veículo": "XYZ-5678",
                "Valor Total": 2300.75,
                "Data Emissão": "2025-01-16",
                "Data Baixa": "",
                "Número Fatura": "",
                "Data Inclusão Fatura": "",
                "Data Envio Processo": "",
                "Primeiro Envio": "",
                "Data RQ/TMC": "",
                "Data Atesto": "",
                "Envio Final": "",
                "Observação": "CTE ainda em processo",
                "Origem dos Dados": "CSV"
            }
        ]
        
        # Criar DataFrame com os dados de exemplo
        df = pd.DataFrame(dados_exemplo)
        
        # Garantir que todas as colunas existam na ordem correta
        df = df.reindex(columns=headers, fill_value="")
        
        buffer = BytesIO()
        
        # Criar Excel com formatação
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template_CTEs', index=False)
            
            # Obter a planilha para formatação
            workbook = writer.book
            worksheet = writer.sheets['Template_CTEs']
            
            # Formatação das colunas
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Cabeçalho em negrito e com fundo
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Ajustar largura das colunas
            column_widths = [
                12,  # Número CTE
                25,  # Destinatário
                12,  # Placa Veículo
                12,  # Valor Total
                12,  # Data Emissão
                12,  # Data Baixa
                15,  # Número Fatura
                18,  # Data Inclusão Fatura
                18,  # Data Envio Processo
                15,  # Primeiro Envio
                12,  # Data RQ/TMC
                12,  # Data Atesto
                12,  # Envio Final
                30,  # Observação
                15   # Origem dos Dados
            ]
            
            for i, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[chr(64 + i)].width = width
        
        buffer.seek(0)
        return buffer

    @staticmethod
    def _limpar_valor_monetario(valor):
        """Limpa e converte valores monetários brasileiros para decimal."""
        if pd.isna(valor) or valor in (None, ""):
            return None
            
        valor_str = str(valor).strip()
        if not valor_str:
            return None
            
        # Remove símbolos monetários e espaços
        valor_str = valor_str.replace("R$", "").replace("$", "").strip()
        
        # Se contém vírgula e ponto, assume formato brasileiro (1.234,56)
        if "," in valor_str and "." in valor_str:
            # Remove pontos (separadores de milhares) e troca vírgula por ponto
            valor_str = valor_str.replace(".", "").replace(",", ".")
        elif "," in valor_str and valor_str.count(",") == 1:
            # Apenas vírgula - pode ser decimal brasileiro
            if len(valor_str.split(",")[1]) <= 2:  # máximo 2 dígitos após vírgula
                valor_str = valor_str.replace(",", ".")
        
        try:
            return float(valor_str)
        except ValueError:
            return None

    @staticmethod  
    def _limpar_data(data_str):
        """Limpa e converte datas para formato ISO (YYYY-MM-DD)."""
        if pd.isna(data_str) or data_str in (None, ""):
            return None
            
        data_str = str(data_str).strip()
        if not data_str:
            return None
            
        # Já está no formato ISO
        if len(data_str) == 10 and data_str.count("-") == 2:
            return data_str
            
        # Formato brasileiro DD/MM/YYYY ou DD/MM/YY
        if "/" in data_str:
            partes = data_str.split("/")
            if len(partes) == 3:
                dia, mes, ano = partes
                # Converte ano de 2 dígitos para 4
                if len(ano) == 2:
                    ano_int = int(ano)
                    if ano_int < 50:  # assume 20xx
                        ano = f"20{ano}"
                    else:  # assume 19xx
                        ano = f"19{ano}"
                        
                return f"{ano}-{mes.zfill(2)}-{dia.zfill(2)}"
        
        return data_str

    @staticmethod
    def _normalizar_dados_linha(dados):
        """Normaliza uma linha de dados importados."""
        dados_limpos = {}
        
        for campo, valor in dados.items():
            if campo == "numero_cte":
                # Limpa número do CTE
                if pd.notna(valor) and valor != "":
                    try:
                        dados_limpos[campo] = int(float(str(valor).strip()))
                    except ValueError:
                        dados_limpos[campo] = None
                else:
                    dados_limpos[campo] = None
                    
            elif campo == "valor_total":
                # Limpa valor monetário
                dados_limpos[campo] = AtualizacaoService._limpar_valor_monetario(valor)
                
            elif campo.startswith("data_"):
                # Limpa datas
                dados_limpos[campo] = AtualizacaoService._limpar_data(valor)
                
            elif campo in ["destinatario_nome", "veiculo_placa", "numero_fatura", "observacao", "origem_dados"]:
                # Campos de texto - apenas limpa espaços
                if pd.notna(valor) and valor != "":
                    dados_limpos[campo] = str(valor).strip()
                else:
                    dados_limpos[campo] = None
            else:
                # Outros campos
                dados_limpos[campo] = valor if pd.notna(valor) and valor != "" else None
                
        return dados_limpos
