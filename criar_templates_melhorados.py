#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from io import BytesIO

def criar_template_csv_melhorado():
    """Cria template CSV com formatação correta e instruções."""
    headers = [
        "Número CTE",
        "Destinatário", 
        "Placa Veículo",
        "Valor Total",
        "Data Emissão",
        "Data Baixa",
        "Número Fatura",
        "Data Inclusão Fatura",
        "Data Envio Processo", 
        "Primeiro Envio",
        "Data RQ/TMC",
        "Data Atesto",
        "Envio Final",
        "Observação",
        "Origem dos Dados"
    ]
    
    # Linha de cabeçalho
    csv_content = ";".join(headers) + "\n"
    
    # Comentários explicativos sobre formato esperado
    csv_content += ";".join([
        "# FORMATOS ESPERADOS:",
        "Razão Social Completa",
        "ABC-1234",
        "1500.50 (decimal com ponto)",
        "AAAA-MM-DD (2025-01-15)",
        "AAAA-MM-DD ou vazio",
        "Número/Código da Fatura",
        "AAAA-MM-DD ou vazio",
        "AAAA-MM-DD ou vazio",
        "AAAA-MM-DD ou vazio",
        "AAAA-MM-DD ou vazio",
        "AAAA-MM-DD ou vazio",
        "AAAA-MM-DD ou vazio",
        "Texto livre",
        "Sistema/CSV/Manual"
    ]) + "\n"
    
    # Linha de exemplo com dados de amostra corretamente formatados
    exemplo = [
        "12345",
        "Cliente Exemplo Ltda",
        "ABC-1234",
        "1500.50",
        "2025-01-15",
        "2025-01-20", 
        "FAT001",
        "2025-01-22",
        "2025-01-25",
        "2025-01-26",
        "2025-01-28",
        "2025-01-30",
        "2025-02-01",
        "Exemplo de observação do CTE",
        "Sistema"
    ]
    csv_content += ";".join(exemplo) + "\n"
    
    # Exemplo adicional com campos opcionais vazios
    exemplo2 = [
        "12346",
        "Outro Cliente SA",
        "XYZ-5678",
        "2300.75",
        "2025-01-16",
        "",  # Data Baixa vazia
        "",  # Número Fatura vazio
        "",  # Data Inclusão Fatura vazia
        "",  # Data Envio Processo vazia
        "",  # Primeiro Envio vazio
        "",  # Data RQ/TMC vazia
        "",  # Data Atesto vazia
        "",  # Envio Final vazio
        "CTE ainda em processo",
        "CSV"
    ]
    csv_content += ";".join(exemplo2) + "\n"
    
    return csv_content

def criar_template_excel_melhorado():
    """Cria template Excel com formatação correta e instruções."""
    headers = [
        "Número CTE",
        "Destinatário", 
        "Placa Veículo",
        "Valor Total",
        "Data Emissão",
        "Data Baixa",
        "Número Fatura",
        "Data Inclusão Fatura",
        "Data Envio Processo", 
        "Primeiro Envio",
        "Data RQ/TMC",
        "Data Atesto",
        "Envio Final",
        "Observação",
        "Origem dos Dados"
    ]
    
    # Linha de instruções sobre formatos
    instrucoes = {
        "Número CTE": "Número inteiro",
        "Destinatário": "Razão Social Completa",
        "Placa Veículo": "ABC-1234",
        "Valor Total": "1500.50 (decimal)",
        "Data Emissão": "AAAA-MM-DD",
        "Data Baixa": "AAAA-MM-DD ou vazio",
        "Número Fatura": "Código/Número",
        "Data Inclusão Fatura": "AAAA-MM-DD ou vazio",
        "Data Envio Processo": "AAAA-MM-DD ou vazio",
        "Primeiro Envio": "AAAA-MM-DD ou vazio",
        "Data RQ/TMC": "AAAA-MM-DD ou vazio",
        "Data Atesto": "AAAA-MM-DD ou vazio",
        "Envio Final": "AAAA-MM-DD ou vazio",
        "Observação": "Texto livre",
        "Origem dos Dados": "Sistema/CSV/Manual"
    }
    
    # Dados de exemplo
    dados_exemplo = [
        instrucoes,  # Primeira linha com instruções
        {
            "Número CTE": 12345,
            "Destinatário": "Cliente Exemplo Ltda",
            "Placa Veículo": "ABC-1234",
            "Valor Total": 1500.50,
            "Data Emissão": "2025-01-15",
            "Data Baixa": "2025-01-20",
            "Número Fatura": "FAT001",
            "Data Inclusão Fatura": "2025-01-22",
            "Data Envio Processo": "2025-01-25",
            "Primeiro Envio": "2025-01-26",
            "Data RQ/TMC": "2025-01-28",
            "Data Atesto": "2025-01-30",
            "Envio Final": "2025-02-01",
            "Observação": "Exemplo de observação do CTE",
            "Origem dos Dados": "Sistema"
        },
        {
            "Número CTE": 12346,
            "Destinatário": "Outro Cliente SA",
            "Placa Veículo": "XYZ-5678",
            "Valor Total": 2300.75,
            "Data Emissão": "2025-01-16",
            "Data Baixa": "",
            "Número Fatura": "",
            "Data Inclusão Fatura": "",
            "Data Envio Processo": "",
            "Primeiro Envio": "",
            "Data RQ/TMC": "",
            "Data Atesto": "",
            "Envio Final": "",
            "Observação": "CTE ainda em processo",
            "Origem dos Dados": "CSV"
        }
    ]
    
    # Criar DataFrame com os dados de exemplo
    df = pd.DataFrame(dados_exemplo)
    
    # Garantir que todas as colunas existam na ordem correta
    df = df.reindex(columns=headers, fill_value="")
    
    buffer = BytesIO()
    
    # Criar Excel com formatação
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Template_CTEs', index=False)
        
        # Obter a planilha para formatação
        workbook = writer.book
        worksheet = writer.sheets['Template_CTEs']
        
        # Formatação das colunas
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Cabeçalho em negrito e com fundo
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Linha de instruções em azul claro
        instruction_font = Font(bold=True, color="000080")
        instruction_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            # Cabeçalho
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
            # Linha de instruções
            instruction_cell = worksheet.cell(row=2, column=col_num)
            instruction_cell.font = instruction_font
            instruction_cell.fill = instruction_fill
            instruction_cell.alignment = Alignment(horizontal="center")
        
        # Ajustar largura das colunas
        column_widths = [
            12,  # Número CTE
            25,  # Destinatário
            12,  # Placa Veículo
            15,  # Valor Total
            15,  # Data Emissão
            15,  # Data Baixa
            15,  # Número Fatura
            18,  # Data Inclusão Fatura
            18,  # Data Envio Processo
            15,  # Primeiro Envio
            15,  # Data RQ/TMC
            15,  # Data Atesto
            15,  # Envio Final
            35,  # Observação
            15   # Origem dos Dados
        ]
        
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + i)].width = width
    
    buffer.seek(0)
    return buffer

if __name__ == "__main__":
    print("=== CRIANDO TEMPLATES MELHORADOS ===")
    
    print("\n1. Criando CSV melhorado...")
    csv_content = criar_template_csv_melhorado()
    with open("template_cte_melhorado.csv", "w", encoding="utf-8-sig") as f:
        f.write(csv_content)
    print(f"   ✅ Salvo como template_cte_melhorado.csv ({len(csv_content)} chars)")
    
    print("\n2. Criando Excel melhorado...")
    excel_buffer = criar_template_excel_melhorado()
    with open("template_cte_melhorado.xlsx", "wb") as f:
        f.write(excel_buffer.getvalue())
    print(f"   ✅ Salvo como template_cte_melhorado.xlsx ({excel_buffer.getbuffer().nbytes} bytes)")
    
    print("\n3. Testando leitura dos templates...")
    
    # Teste CSV (pular linha de instruções)
    df_csv = pd.read_csv("template_cte_melhorado.csv", sep=";", skiprows=1)
    print(f"   📊 CSV: {df_csv.shape[0]} linhas, {df_csv.shape[1]} colunas")
    
    # Teste Excel (pular linha de instruções)
    df_excel = pd.read_excel("template_cte_melhorado.xlsx", skiprows=1)
    print(f"   📊 Excel: {df_excel.shape[0]} linhas, {df_excel.shape[1]} colunas")
    
    print("\n🎉 Templates melhorados criados com sucesso!")
    print("\n📋 INSTRUÇÕES PARA USO:")
    print("   - Valores monetários: use formato decimal com ponto (1500.50)")
    print("   - Datas: use formato ISO (AAAA-MM-DD)")
    print("   - Campos opcionais: deixe em branco se não aplicável")
    print("   - Ao importar: remova ou pule a linha de instruções")
