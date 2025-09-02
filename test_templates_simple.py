#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from io import BytesIO

class AtualizacaoServiceSimple:
    """Versão simplificada para testar templates."""
    
    @staticmethod
    def template_csv() -> str:
        """Gera template CSV completo para download."""
        headers = [
            "Número CTE",
            "Destinatário", 
            "Placa Veículo",
            "Valor Total",
            "Data Emissão",
            "Data Baixa",
            "Número Fatura",
            "Data Inclusão Fatura",
            "Data Envio Processo", 
            "Primeiro Envio",
            "Data RQ/TMC",
            "Data Atesto",
            "Envio Final",
            "Observação",
            "Origem dos Dados"
        ]
        
        # Linha de cabeçalho
        csv_content = ";".join(headers) + "\n"
        
        # Linha de exemplo com dados de amostra
        exemplo = [
            "12345",
            "Cliente Exemplo Ltda",
            "ABC-1234",
            "1500.50",
            "2025-01-15",
            "2025-01-20", 
            "FAT001",
            "2025-01-22",
            "2025-01-25",
            "2025-01-26",
            "2025-01-28",
            "2025-01-30",
            "2025-02-01",
            "Exemplo de observação do CTE",
            "Sistema"
        ]
        csv_content += ";".join(exemplo) + "\n"
        
        return csv_content

    @staticmethod
    def template_excel() -> BytesIO:
        """Gera template Excel completo para download."""
        headers = [
            "Número CTE",
            "Destinatário", 
            "Placa Veículo",
            "Valor Total",
            "Data Emissão",
            "Data Baixa",
            "Número Fatura",
            "Data Inclusão Fatura",
            "Data Envio Processo", 
            "Primeiro Envio",
            "Data RQ/TMC",
            "Data Atesto",
            "Envio Final",
            "Observação",
            "Origem dos Dados"
        ]
        
        # Dados de exemplo
        dados_exemplo = [
            {
                "Número CTE": 12345,
                "Destinatário": "Cliente Exemplo Ltda",
                "Placa Veículo": "ABC-1234",
                "Valor Total": 1500.50,
                "Data Emissão": "2025-01-15",
                "Data Baixa": "2025-01-20",
                "Número Fatura": "FAT001",
                "Data Inclusão Fatura": "2025-01-22",
                "Data Envio Processo": "2025-01-25",
                "Primeiro Envio": "2025-01-26",
                "Data RQ/TMC": "2025-01-28",
                "Data Atesto": "2025-01-30",
                "Envio Final": "2025-02-01",
                "Observação": "Exemplo de observação do CTE",
                "Origem dos Dados": "Sistema"
            },
            {
                "Número CTE": 12346,
                "Destinatário": "Outro Cliente SA",
                "Placa Veículo": "XYZ-5678",
                "Valor Total": 2300.75,
                "Data Emissão": "2025-01-16",
                "Data Baixa": "",
                "Número Fatura": "",
                "Data Inclusão Fatura": "",
                "Data Envio Processo": "",
                "Primeiro Envio": "",
                "Data RQ/TMC": "",
                "Data Atesto": "",
                "Envio Final": "",
                "Observação": "CTE ainda em processo",
                "Origem dos Dados": "Sistema"
            }
        ]
        
        # Criar DataFrame com os dados de exemplo
        df = pd.DataFrame(dados_exemplo)
        
        # Garantir que todas as colunas existam na ordem correta
        df = df.reindex(columns=headers, fill_value="")
        
        buffer = BytesIO()
        
        # Criar Excel com formatação
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template_CTEs', index=False)
            
            # Obter a planilha para formatação
            workbook = writer.book
            worksheet = writer.sheets['Template_CTEs']
            
            # Formatação das colunas
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Cabeçalho em negrito e com fundo
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Ajustar largura das colunas
            column_widths = [
                12,  # Número CTE
                25,  # Destinatário
                12,  # Placa Veículo
                12,  # Valor Total
                12,  # Data Emissão
                12,  # Data Baixa
                15,  # Número Fatura
                18,  # Data Inclusão Fatura
                18,  # Data Envio Processo
                15,  # Primeiro Envio
                12,  # Data RQ/TMC
                12,  # Data Atesto
                12,  # Envio Final
                30,  # Observação
                15   # Origem dos Dados
            ]
            
            for i, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[chr(64 + i)].width = width
        
        buffer.seek(0)
        return buffer

if __name__ == "__main__":
    print("Testing CSV template...")
    csv_content = AtualizacaoServiceSimple.template_csv()
    print(f"CSV template length: {len(csv_content)}")
    
    print("\nCSV content preview:")
    lines = csv_content.strip().split('\n')
    for i, line in enumerate(lines[:3]):
        print(f"Line {i+1}: {line}")
    
    print("\nTesting Excel template...")
    excel_buffer = AtualizacaoServiceSimple.template_excel()
    print(f"Excel buffer size: {excel_buffer.getbuffer().nbytes} bytes")
    
    # Test saving to file
    print("\nSaving templates to files...")
    with open("template_test.csv", "w", encoding="utf-8-sig") as f:
        f.write(csv_content)
    print("CSV template saved to template_test.csv")
    
    with open("template_test.xlsx", "wb") as f:
        f.write(excel_buffer.getvalue())
    print("Excel template saved to template_test.xlsx")
    
    print("\nAll tests passed!")
